from decimal import Decimal
from io import BytesIO, StringIO, TextIOWrapper
import csv

from rest_framework import permissions, viewsets
from rest_framework.exceptions import PermissionDenied
from rest_framework.decorators import action
from rest_framework.response import Response

from .models_coef import Coeficiente
from .serializers_coef import CoeficienteSerializer


class CoeficienteViewSet(viewsets.ModelViewSet):
    queryset = Coeficiente.objects.all()
    serializer_class = CoeficienteSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_permissions(self):  # pragma: no cover - simple auth tweak
        if self.action in ["list", "retrieve"]:
            return [permissions.IsAuthenticated()]
        return super().get_permissions()

    def _ensure_superuser(self):
        user = getattr(self.request, "user", None)
        if not (user and user.is_authenticated and user.is_superuser):
            raise PermissionDenied("Apenas superadministradores podem alterar coeficientes.")

    def create(self, request, *args, **kwargs):  # pragma: no cover - simple guard
        self._ensure_superuser()
        return super().create(request, *args, **kwargs)

    def update(self, request, *args, **kwargs):  # pragma: no cover - simple guard
        self._ensure_superuser()
        return super().update(request, *args, **kwargs)

    def partial_update(self, request, *args, **kwargs):  # pragma: no cover - simple guard
        self._ensure_superuser()
        return super().partial_update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):  # pragma: no cover - simple guard
        self._ensure_superuser()
        return super().destroy(request, *args, **kwargs)

    @action(detail=False, methods=["post"], url_path="import")
    def import_table(self, request):
        """Importa tabela de coeficientes via XLSX ou CSV."""
        self._ensure_superuser()
        uploaded = request.FILES.get("file")
        inline_csv = request.data.get("csv")
        processed = 0

        if uploaded and uploaded.name.lower().endswith(".xlsx"):
            try:
                from openpyxl import load_workbook  # type: ignore
            except Exception:  # pragma: no cover - depends on optional dependency
                return Response({"detail": "Instale openpyxl"}, status=500)

            wb = load_workbook(filename=BytesIO(uploaded.read()))
            ws = wb.active
            header = [
                str(cell.value).strip().lower() if cell.value is not None else ""
                for cell in next(ws.iter_rows(min_row=1, max_row=1))
            ]
            try:
                idx = {name: header.index(name) for name in ["banco", "parcelas", "coeficiente"]}
            except ValueError:
                return Response({"detail": "Planilha precisa das colunas banco, parcelas, coeficiente."}, status=422)

            for row in ws.iter_rows(min_row=2):
                banco = str(row[idx["banco"]].value or "").strip()
                parcelas = int(row[idx["parcelas"]].value or 0)
                coef = Decimal(str(row[idx["coeficiente"]].value or "0"))
                if not banco or parcelas <= 0 or coef <= 0:
                    continue
                Coeficiente.objects.update_or_create(
                    banco=banco,
                    parcelas=parcelas,
                    defaults={"coeficiente": coef},
                )
                processed += 1
            return Response({"ok": True, "rows": processed, "format": "xlsx"})

        if uploaded:
            inline_csv = TextIOWrapper(uploaded.file, encoding="utf-8").read()

        if inline_csv:
            reader = csv.DictReader(StringIO(inline_csv))
            for row in reader:
                banco = (row.get("banco") or "").strip()
                parcelas = int((row.get("parcelas") or "0").strip() or 0)
                coef = Decimal(str(row.get("coeficiente") or "0"))
                if not banco or parcelas <= 0 or coef <= 0:
                    continue
                Coeficiente.objects.update_or_create(
                    banco=banco,
                    parcelas=parcelas,
                    defaults={"coeficiente": coef},
                )
                processed += 1
            return Response({"ok": True, "rows": processed, "format": "csv"})

        return Response({"detail": "Envie 'file' (.xlsx/.csv) ou 'csv' (texto)."}, status=400)
