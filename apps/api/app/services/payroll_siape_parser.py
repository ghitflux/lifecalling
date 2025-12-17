"""
Parser para arquivos SIAPE (.xlsx).

Este parser processa planilhas Excel com dados de empréstimos consignados
de funcionários públicos federais (SIAPE - Sistema Integrado de Administração
de Recursos Humanos).
"""

import logging
from datetime import datetime
from decimal import Decimal
from typing import Dict, List, Tuple
from openpyxl import load_workbook
import re

# Configure logging for import diagnostics
logger = logging.getLogger(__name__)


def normalize_cpf(cpf: str) -> str:
    """Remove formatação do CPF, mantendo apenas dígitos"""
    if cpf is None:
        return ""

    # Se vier como número (int/float), converte para int primeiro para evitar ".0"
    if isinstance(cpf, (int, float, Decimal)):
        try:
            cpf_str = str(int(cpf))
        except Exception:
            cpf_str = str(cpf)
    else:
        cpf_str = str(cpf).strip()

    return "".join(ch for ch in cpf_str if ch.isdigit())


def normalize_phone(phone: str) -> str:
    """Remove formatação do telefone, mantendo apenas dígitos"""
    if phone is None:
        return ""

    # Se vier como número (int/float), converte para int para remover casas decimais .0
    if isinstance(phone, (int, float, Decimal)):
        try:
            phone_str = str(int(phone))
        except Exception:
            phone_str = str(phone)
    else:
        phone_str = str(phone).strip()

    # Remove todos os caracteres não numéricos
    digits = "".join(ch for ch in phone_str if ch.isdigit())
    # Remove código de país se tiver (55)
    if digits.startswith("55") and len(digits) >= 12:
        digits = digits[2:]
    return digits


def normalize_currency(value) -> Decimal:
    """Converte valores como "1.234,56", "30.00", "17,94" ou floats em Decimal."""
    if value is None:
        return Decimal("0")

    # Se já for número (float/int), retornar direto
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value))

    # Se for string, processar
    val = str(value).strip()

    # Se contém vírgula e ponto, assume formato brasileiro: remove pontos de milhar e troca vírgula por ponto
    if "," in val and "." in val:
        normalized = val.replace(".", "").replace(",", ".")
    # Se contém apenas vírgula, vírgula é o separador decimal
    elif "," in val:
        normalized = val.replace(",", ".")
    # Se contém apenas ponto, ponto é o separador decimal
    else:
        normalized = val

    try:
        return Decimal(normalized)
    except Exception:
        logger.warning(f"Erro ao converter valor '{value}' para Decimal")
        return Decimal("0")


def normalize_date(date_value) -> str:
    """Converte datas Excel em formato DD/MM/YYYY"""
    if date_value is None:
        return ""

    # Se já for datetime, formatar
    if isinstance(date_value, datetime):
        return date_value.strftime("%d/%m/%Y")

    # Se for string, retornar como está
    if isinstance(date_value, str):
        return date_value.strip()

    return str(date_value)


def normalize_cep(cep) -> str:
    """Normaliza CEP removendo formatação"""
    if cep is None:
        return ""

    # Tratar números (float/int/Decimal) para remover casas decimais e não perder zeros à esquerda
    if isinstance(cep, (int, float, Decimal)):
        try:
            cep_str = str(int(cep))
        except Exception:
            cep_str = str(cep)
    else:
        cep_str = str(cep).strip()

    digits = "".join(ch for ch in cep_str if ch.isdigit())
    # Garantir no máximo 8 dígitos para caber no campo VARCHAR(8)
    if len(digits) > 8:
        digits = digits[:8]
    return digits


def parse_siape_file(file_path: str) -> Tuple[Dict, List[Dict], Dict]:
    """
    Função principal para fazer parse completo de um arquivo SIAPE (.xlsx).

    Args:
        file_path: Caminho do arquivo Excel

    Returns:
        Tuple contendo:
        - meta: Metadados do arquivo
        - lines: Lista de linhas processadas
        - stats: Estatísticas do processamento

    Raises:
        ValueError: Se o formato do arquivo for inválido
    """
    logger.info(f"Iniciando parse do arquivo SIAPE: {file_path}")

    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active
    except Exception as e:
        raise ValueError(f"Erro ao abrir arquivo Excel: {str(e)}")

    # Metadados (fixos para SIAPE)
    meta = {
        "entity_code": "SIAPE",
        "entity_name": "Sistema Integrado de Administração de Recursos Humanos",
        "ref_month": datetime.now().month,
        "ref_year": datetime.now().year,
        "generated_at": datetime.now()
    }

    # Ler cabeçalhos da primeira linha
    headers_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(headers_row)]

    logger.info(f"Cabeçalhos encontrados: {headers}")

    # Mapear índices de colunas importantes
    column_map = {}
    for i, header in enumerate(headers):
        header_lower = header.lower()
        if 'cpf' in header_lower and 'cpf' not in column_map:
            column_map['cpf'] = i
        elif 'nome' in header_lower and 'nome' not in column_map:
            column_map['nome'] = i
        elif 'nascimento' in header_lower and 'nascimento' not in column_map:
            column_map['nascimento'] = i
        elif 'idade' in header_lower and 'idade' not in column_map:
            column_map['idade'] = i
        elif 'matricula' in header_lower and 'matricula' not in column_map:
            column_map['matricula'] = i
        elif 'banco' in header_lower and 'emprestimo' in header_lower and 'banco_emprestimo' not in column_map:
            column_map['banco_emprestimo'] = i
        elif header_lower == 'prazo' and 'prazo' not in column_map:
            column_map['prazo'] = i
        elif 'prazo total' in header_lower and 'prazo_total' not in column_map:
            column_map['prazo_total'] = i
        elif 'parcelas pagas' in header_lower and 'parcelas_pagas' not in column_map:
            column_map['parcelas_pagas'] = i
        elif 'valor' in header_lower and 'parcela' in header_lower and 'valor_parcela' not in column_map:
            column_map['valor_parcela'] = i
        elif 'saldo' in header_lower and 'saldo_devedor' not in column_map:
            column_map['saldo_devedor'] = i
        elif ('fone' in header_lower or 'telefone' in header_lower) and 'telefone' not in column_map:
            column_map['telefone'] = i
        elif ('e-mail' in header_lower or 'email' in header_lower) and 'email' not in column_map:
            column_map['email'] = i
        elif 'cidade' in header_lower and 'cidade' not in column_map:
            column_map['cidade'] = i
        elif 'bairro' in header_lower and 'bairro' not in column_map:
            column_map['bairro'] = i
        elif 'cep' in header_lower and 'cep' not in column_map:
            column_map['cep'] = i
        elif ('endereco' in header_lower or 'endereço' in header_lower) and 'endereco' not in column_map:
            column_map['endereco'] = i

    logger.info(f"Mapeamento de colunas: {column_map}")

    # Validar colunas obrigatórias
    required_columns = ['cpf', 'nome', 'matricula', 'banco_emprestimo']
    missing_columns = [col for col in required_columns if col not in column_map]
    if missing_columns:
        raise ValueError(f"Colunas obrigatórias não encontradas: {missing_columns}")

    # Parse das linhas de dados
    lines = []
    stats = {
        "total_lines_processed": 0,
        "valid_lines": 0,
        "invalid_lines": 0,
        "unique_cpfs": set(),
        "errors_sample": []
    }

    # Processar linhas (começando da linha 2)
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
        stats["total_lines_processed"] += 1

        try:
            # Extrair CPF (obrigatório)
            cpf_raw = row[column_map['cpf']] if 'cpf' in column_map else None
            cpf = normalize_cpf(cpf_raw)

            if not cpf or len(cpf) != 11:
                logger.warning(f"Linha {row_idx}: CPF inválido '{cpf_raw}'")
                stats["invalid_lines"] += 1
                continue

            # Extrair dados básicos
            nome = str(row[column_map['nome']]).strip() if 'nome' in column_map and row[column_map['nome']] else ""
            matricula = str(row[column_map['matricula']]).strip() if 'matricula' in column_map and row[column_map['matricula']] else ""
            banco_emprestimo = str(row[column_map['banco_emprestimo']]).strip() if 'banco_emprestimo' in column_map and row[column_map['banco_emprestimo']] else ""

            # Validar campos obrigatórios
            if not nome or not matricula:
                logger.warning(f"Linha {row_idx}: Nome ou matrícula vazio")
                stats["invalid_lines"] += 1
                continue

            # Extrair dados opcionais
            nascimento = normalize_date(row[column_map['nascimento']]) if 'nascimento' in column_map else ""
            idade = int(row[column_map['idade']]) if 'idade' in column_map and row[column_map['idade']] else None

            # Dados do empréstimo
            prazo_restante = int(row[column_map['prazo']]) if 'prazo' in column_map and row[column_map['prazo']] else None
            prazo_total = int(row[column_map['prazo_total']]) if 'prazo_total' in column_map and row[column_map['prazo_total']] else None
            parcelas_pagas = int(row[column_map['parcelas_pagas']]) if 'parcelas_pagas' in column_map and row[column_map['parcelas_pagas']] else None

            # Se não tem parcelas_pagas mas tem prazo_total e prazo_restante, calcular
            if parcelas_pagas is None and prazo_total and prazo_restante:
                parcelas_pagas = prazo_total - prazo_restante

            valor_parcela = normalize_currency(row[column_map['valor_parcela']]) if 'valor_parcela' in column_map else Decimal("0")
            saldo_devedor = normalize_currency(row[column_map['saldo_devedor']]) if 'saldo_devedor' in column_map else Decimal("0")

            # Dados de contato
            telefone = normalize_phone(row[column_map['telefone']]) if 'telefone' in column_map else ""
            email = str(row[column_map['email']]).strip() if 'email' in column_map and row[column_map['email']] else ""

            # Dados de endereço
            cidade = str(row[column_map['cidade']]).strip() if 'cidade' in column_map and row[column_map['cidade']] else ""
            bairro = str(row[column_map['bairro']]).strip() if 'bairro' in column_map and row[column_map['bairro']] else ""
            cep = normalize_cep(row[column_map['cep']]) if 'cep' in column_map else ""
            endereco = str(row[column_map['endereco']]).strip() if 'endereco' in column_map and row[column_map['endereco']] else ""

            # Criar registro da linha
            line_data = {
                # Metadados
                "line_number": row_idx,
                "entity_code": meta["entity_code"],
                "entity_name": meta["entity_name"],
                "ref_month": meta["ref_month"],
                "ref_year": meta["ref_year"],

                # Dados do cliente
                "cpf": cpf,
                "nome": nome.upper(),  # Normalizar para maiúsculas
                "matricula": matricula,
                "nascimento": nascimento,
                "idade": idade,

                # Dados do empréstimo
                "banco_emprestimo": banco_emprestimo,
                "prazo_restante": prazo_restante,
                "prazo_total": prazo_total or prazo_restante,  # Se não tem total, usar restante
                "parcelas_pagas": parcelas_pagas or 0,
                "valor_parcela": valor_parcela,
                "saldo_devedor": saldo_devedor,

                # Status SIAPE (sempre "ativo" para importação)
                "status_code": "A",  # A = Ativo
                "status_description": "Empréstimo Ativo SIAPE",

                # Dados de contato
                "telefone": telefone,
                "email": email,

                # Dados de endereço
                "cidade": cidade,
                "bairro": bairro,
                "cep": cep,
                "endereco": endereco,

                # Financiamento code (usar hash do banco para simular FIN)
                "financiamento_code": f"SIAPE_{abs(hash(banco_emprestimo)) % 10000:04d}",
                "orgao_pagamento": "SIAPE",
                "orgao_pagamento_nome": "Sistema Integrado de Administração de RH",
            }

            lines.append(line_data)
            stats["valid_lines"] += 1
            stats["unique_cpfs"].add(cpf)

        except Exception as e:
            stats["invalid_lines"] += 1
            error_msg = f"Linha {row_idx}: {str(e)}"
            logger.error(f"Erro ao processar {error_msg}")
            if len(stats["errors_sample"]) < 10:
                stats["errors_sample"].append(error_msg)

    # Fechar workbook
    wb.close()

    # Estatísticas finais
    final_stats = {
        "total_lines": len(lines),
        "unique_clients": stats["valid_lines"],  # Cada linha é um cliente
        "unique_cpfs": len(stats["unique_cpfs"]),
        "entity": f"{meta['entity_code']}-{meta['entity_name']}",
        "reference": f"{meta['ref_month']:02d}/{meta['ref_year']}",
        "processed_lines": stats["valid_lines"],
        "error_lines": stats["invalid_lines"]
    }

    logger.info(f"Parse concluído: {final_stats['total_lines']} linhas válidas, "
                f"{final_stats['error_lines']} inválidas, "
                f"{final_stats['unique_cpfs']} CPFs únicos")

    return meta, lines, final_stats


def validate_siape_content(file_path: str) -> List[str]:
    """
    Validação do arquivo SIAPE.

    Args:
        file_path: Caminho do arquivo Excel

    Returns:
        Lista de strings com erros encontrados (vazia se válido)
    """
    errors = []

    try:
        wb = load_workbook(file_path, read_only=True)
        ws = wb.active

        # Contar linhas de dados (ignorando o header) pois ws.max_row pode vir None
        data_rows = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(cell not in (None, "") for cell in row):
                data_rows += 1
        total_rows = data_rows + 1  # +1 pelo header

        # Verificar se tem dados
        if data_rows == 0:
            errors.append("Arquivo não contém dados")
            return errors

        # Ler cabeçalhos
        headers_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = [str(h).lower().strip() if h else "" for h in headers_row]

        # Verificar colunas obrigatórias
        required = ['cpf', 'nome', 'matricula']
        for req in required:
            if not any(req in h for h in headers):
                errors.append(f"Coluna obrigatória '{req.upper()}' não encontrada")

        wb.close()

    except Exception as e:
        errors.append(f"Erro ao validar arquivo: {str(e)}")

    return errors


def get_file_preview(file_path: str, max_lines: int = 5) -> Dict:
    """
    Retorna uma prévia do arquivo SIAPE para diagnóstico.

    Args:
        file_path: Caminho do arquivo
        max_lines: Número máximo de linhas a mostrar

    Returns:
        Dict com prévia do arquivo
    """
    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active

        # Contar linhas de dados (ignorando o header)
        data_rows = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(cell not in (None, "") for cell in row):
                data_rows += 1
        total_rows = data_rows + 1  # +1 pelo header

        # Ler cabeçalhos
        headers_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(headers_row)]

        # Ler algumas linhas de exemplo
        sample_lines = []
        for i, row in enumerate(ws.iter_rows(min_row=2, max_row=max_lines+1, values_only=True)):
            if i >= max_lines:
                break

            # Extrair CPF e Nome
            cpf_idx = next((i for i, h in enumerate(headers) if 'cpf' in h.lower()), None)
            nome_idx = next((i for i, h in enumerate(headers) if 'nome' in h.lower()), None)

            sample_lines.append({
                "cpf": normalize_cpf(row[cpf_idx]) if cpf_idx is not None else "",
                "nome": str(row[nome_idx]).strip() if nome_idx is not None and row[nome_idx] else "",
            })

        wb.close()

        return {
            "total_lines": max(total_rows - 1, 0),  # -1 para excluir header
            "headers": headers,
            "sample_data_lines": sample_lines,
            "estimated_records": max(total_rows - 1, 0)
        }

    except Exception as e:
        return {
            "error": str(e),
            "total_lines": 0,
            "headers": [],
            "sample_data_lines": [],
            "estimated_records": 0
        }
