from openpyxl import load_workbook

wb = load_workbook('SIAPE.xlsx')
ws = wb.active

headers = [cell.value for cell in ws[1]]
print('Colunas:', headers)
print(f'\nTotal de colunas: {len(headers)}')

print('\nPrimeiras 3 linhas de dados:')
for i in range(2, min(5, ws.max_row+1)):
    row = [cell.value for cell in ws[i]]
    print(f'\nLinha {i-1}:')
    for j, (header, value) in enumerate(zip(headers, row)):
        print(f'  {header}: {value}')

print(f'\nTotal de linhas de dados: {ws.max_row - 1}')
